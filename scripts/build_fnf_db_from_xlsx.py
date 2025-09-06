#!/usr/bin/env python3
import json
from openpyxl import load_workbook

XLSX_PATH = 'Simple Sizer FNF with formulas 1.xlsx'
DB_PATH = 'archwest_fnf_database.json'

def pct(cell):
    if cell is None: return None
    try:
        v = float(cell)
        return v if v <= 1.0 else v/100.0
    except:
        s = str(cell).strip().replace('%','')
        try:
            return float(s)/100.0
        except:
            return None

def money(cell):
    if cell is None: return None
    s = str(cell).replace('$','').replace(',','').strip()
    try:
        return float(s)
    except:
        return None

def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    # Find the row where headers include 'Product' and 'Borrower Level'
    header_row = None
    for r in range(1, ws.max_row+1):
        row_vals = [ (ws.cell(r,c).value or '') for c in range(1, ws.max_column+1) ]
        if 'Product' in row_vals and 'Borrower Level' in row_vals:
            header_row = r
            break
    if header_row is None:
        raise SystemExit('Header row not found')

    # Build column map by scanning the next 2 header rows as well
    labels = {}
    for c in range(1, ws.max_column+1):
        parts = []
        for r in (header_row, header_row+1, header_row+2):
            v = ws.cell(r,c).value
            if v: parts.append(str(v).strip())
        label = ' '.join(parts)
        labels[c] = label

    def col(name_substr):
        for c,label in labels.items():
            if name_substr in label:
                return c
        return None

    cols = {
        'product': col('Product'),
        'level': col('Borrower Level'),
        'minExp': col('Min. Experience 36 mos.'),
        'minFico': col('Min. FICO'),
        'loanAmountTier': col('Loan Amount Tier'),
        'minLoan': col('Min. Loan'),
        'maxLoan': col('Max. Loan'),
        'pLTV': col('Purchase LTV'),
        'pLTARV': col('Purchase LTARV'),
        'pLTC': col('Purchase LTC'),
        'rLTV': col('Refinance LF RehabLTV'),
        'rLTARV': col('Refinance LF RehabLTARV'),
        'rLTC': col('Refinance LF RehabLTC'),
        'rate1': col('Tier 1'),
        'rate2': col('Tier 2'),
        'rate3': col('Tier 3'),
    }

    with open(DB_PATH) as f:
        db = json.load(f)

    rows = []
    for r in range(header_row+1, ws.max_row+1):
        def sval(v):
            return '' if v is None else str(v).strip()
        product = sval(ws.cell(r, cols['product']).value) if cols['product'] else ''
        if product != 'FNF':
            continue
        level = sval(ws.cell(r, cols['level']).value) if cols['level'] else ''
        if level not in ('A','B','C','D'):
            continue
        try:
            minExp = int(sval(ws.cell(r, cols['minExp']).value) or 0) if cols['minExp'] else 0
        except:
            minExp = 0
        try:
            minFico = int(sval(ws.cell(r, cols['minFico']).value) or 0) if cols['minFico'] else 0
        except:
            continue
        try:
            tier = int(sval(ws.cell(r, cols['loanAmountTier']).value) or 0) if cols['loanAmountTier'] else 0
        except:
            tier = 0
        minLoan = money(ws.cell(r, cols['minLoan']).value) if cols['minLoan'] else None
        maxLoan = money(ws.cell(r, cols['maxLoan']).value) if cols['maxLoan'] else None

        p = {
            'LTV': pct(ws.cell(r, cols['pLTV']).value) if cols['pLTV'] else None,
            'LTARV': pct(ws.cell(r, cols['pLTARV']).value) if cols['pLTARV'] else None,
            'LTC': pct(ws.cell(r, cols['pLTC']).value) if cols['pLTC'] else None,
        }
        rf = {
            'LTV': pct(ws.cell(r, cols['rLTV']).value) if cols['rLTV'] else None,
            'LTARV': pct(ws.cell(r, cols['rLTARV']).value) if cols['rLTARV'] else None,
            'LTC': pct(ws.cell(r, cols['rLTC']).value) if cols['rLTC'] else None,
        }
        rates = {
            'Tier1': pct(ws.cell(r, cols['rate1']).value) if cols['rate1'] else None,
            'Tier2': pct(ws.cell(r, cols['rate2']).value) if cols['rate2'] else None,
            'Tier3': pct(ws.cell(r, cols['rate3']).value) if cols['rate3'] else None,
        }
        if tier and minLoan and maxLoan:
            rows.append({
                'product':'FNF','borrowerLevel':level,'minExperienceMonths':minExp,'minFico':minFico,
                'loanAmountTier':tier,'minLoan':minLoan,'maxLoan':maxLoan,
                'purchase':p,'refi':rf,'noteRates':rates
            })

    # merge into db (non-null overwrite)
    def k(r): return (r['borrowerLevel'], r['minFico'], r['loanAmountTier'])
    byk = {k(r): r for r in db['products']['FNF']['pricing_rows']}
    for r in rows:
        key = k(r)
        if key not in byk:
            byk[key] = r
            continue
        base = byk[key]
        for blk in ('purchase','refi'):
            for cap in ('LTV','LTARV','LTC'):
                if r[blk].get(cap) is not None:
                    base[blk][cap] = r[blk][cap]
        if r.get('noteRates'):
            base['noteRates'] = r['noteRates']
        base['minLoan'] = r['minLoan'] or base['minLoan']
        base['maxLoan'] = r['maxLoan'] or base['maxLoan']
        base['minExperienceMonths'] = r['minExperienceMonths'] or base.get('minExperienceMonths')

    db['products']['FNF']['pricing_rows'] = sorted(byk.values(), key=lambda x:(x['borrowerLevel'],x['minFico'],x['loanAmountTier']))

    with open(DB_PATH,'w') as f:
        json.dump(db,f,indent=2)

    print('Wrote', len(db['products']['FNF']['pricing_rows']), 'rows to', DB_PATH)

if __name__ == '__main__':
    main()


